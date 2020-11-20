# Handing Excel spreadsheets w/ openpyxl

import openpyxl
import os
os.chdir('c:\\users\\Mason\\documents')
workbook = openpyxl.load_workbook('example.xlsx')
type(workbook)
# <class 'openpyxl.workbook.workbook.Workbook'>
sheet = workbook.get_sheet_by_name('Sheet1')
type(sheet)
# <class 'openpyxl.worksheet.worksheet.Worksheet'>
workbook.get_sheet_names()
# ['Sheet1', 'Sheet2', 'Sheet3']

# to get cell value:
sheet['A1'] #cell A1
# <Cell 'Sheet1'.A1>
cell = sheet['A1']
cell.value
# datetime.datetime(2015, 4, 5, 13, 34, 2)
str(cell.value)
# '2015-04-05 13:34:02'
str(sheet['A1'].value)
# '2015-04-05 13:34:02'
sheet['B1'].value
# 'Apples'
sheet['C1'].value
# 73
str(sheet['C1'].value) # to get string value
# '73' 

sheet.cell(row=1, column=2)
# <Cell 'Sheet1'.B1>
sheet['B1']
# <Cell 'Sheet1'.B1>
# these two are same thing, though former is neater

# iterating over the spreadsheet
for i in range(1,8):
	print(i, sheet.cell(row=i, column=2).value)

	
# 1 Apples
# 2 Cherries
# 3 Pears
# 4 Oranges
# 5 Apples
# 6 Bananas
# 7 Strawberries
