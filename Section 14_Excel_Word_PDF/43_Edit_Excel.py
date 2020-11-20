# Create and modify excel spreadsheets
import openpyxl
wb = openpyxl.Workbook()
wb
# <openpyxl.workbook.workbook.Workbook object at 0x03344580>
wb.get_sheet_names()
# ['Sheet']
sheet = wb.get_sheet_by_name('Sheet')
sheet
# <Worksheet "Sheet">
sheet['A1'].value
sheet['A1'].value == None
# True
sheet['A1'] = 42
sheet['A2'] = 'Hello'

import os
os.chdir('c:\\users\Mason\Documents')
wb.save('example.xlsx')
sheet2 = wb.create_sheet()
wb.get_sheet_names()
# ['Sheet', 'Sheet1']
sheet2.title
# 'Sheet1'
sheet2.title = 'My New Sheet Name'
wb.get_sheet_names()
# ['Sheet', 'My New Sheet Name']
wb.save('example2.xlsx')

wb.create_sheet(index=0, title='My Other Sheet')
# <Worksheet "My Other Sheet">
wb.save('example3.xlsx')



